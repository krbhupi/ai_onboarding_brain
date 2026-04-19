"""ETL service for processing Excel tracker data."""
import hashlib
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from config.settings import get_settings
from config.logging import logger
from src.models.database import CandidateInfo, CandidateTypeMaster
from src.constants.constants import CandidateType

settings = get_settings()


class ETLService:
    """Service for ETL pipeline: Excel to Database."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.excel_path = Path(settings.EXCEL_TRACKER_PATH)

    def compute_row_hash(self, row: Dict[str, Any]) -> str:
        """Compute a unique hash for a row based on key fields."""
        hash_string = "|".join(str(v) for v in row.values() if v is not None)
        return hashlib.sha256(hash_string.encode()).hexdigest()

    def read_excel_sheets(self) -> List[Dict[str, Any]]:
        """Read all sheets from the Excel tracker file."""
        if not self.excel_path.exists():
            logger.error(f"Excel file not found: {self.excel_path}")
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        workbook = openpyxl.load_workbook(self.excel_path, read_only=True)
        all_rows = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = self._extract_headers(sheet)
            rows = self._extract_rows(sheet, headers, sheet_name)
            all_rows.extend(rows)
            logger.info(f"Read {len(rows)} rows from sheet: {sheet_name}")

        workbook.close()
        return all_rows

    def _extract_headers(self, sheet: Worksheet) -> List[str]:
        """Extract column headers from the first row."""
        headers = []
        for cell in sheet[1]:
            header = str(cell.value).strip().lower().replace(" ", "_") if cell.value else ""
            headers.append(header)
        return headers

    def _extract_rows(
        self,
        sheet: Worksheet,
        headers: List[str],
        sheet_name: str
    ) -> List[Dict[str, Any]]:
        """Extract all data rows from a sheet."""
        rows = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue  # Skip empty rows

            row_data = {
                "sheet_name": sheet_name,
                "row_number": row_num,
            }
            for idx, value in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = value

            row_data["row_hash"] = self.compute_row_hash(row_data)
            rows.append(row_data)

        return rows

    def generate_cin(self, offer_release_date: Any, ref_no: Any) -> str:
        """Generate CIN from offer release date and reference number."""
        date_str = ""
        ref_str = ""

        if offer_release_date:
            if isinstance(offer_release_date, datetime):
                date_str = offer_release_date.strftime("%Y%m%d")
            else:
                date_str = str(offer_release_date).replace("-", "")[:8]

        if ref_no:
            ref_str = str(ref_no).replace("-", "")[:8]

        return f"{date_str}_{ref_str}"

    async def get_or_create_candidate_type(
        self,
        type_name: str
    ) -> Optional[int]:
        """Get candidate type ID from name."""
        type_mapping = {
            "fresher": CandidateType.FRESHER,
            "experience": CandidateType.EXPERIENCE,
            "dev_partner": CandidateType.DEV_PARTNER,
            "dev partner": CandidateType.DEV_PARTNER,
        }

        normalized = type_name.lower().strip() if type_name else ""
        return type_mapping.get(normalized)

    async def sync_candidates(self) -> Dict[str, int]:
        """Sync candidate data from Excel to database."""
        rows = self.read_excel_sheets()
        stats = {"created": 0, "updated": 0, "skipped": 0}

        for row in rows:
            row_hash = row.get("row_hash")
            if not row_hash:
                stats["skipped"] += 1
                continue

            # Check if row already exists
            existing = await self._find_candidate_by_hash(row_hash)

            if existing:
                await self._update_candidate(existing, row)
                stats["updated"] += 1
            else:
                await self._create_candidate(row)
                stats["created"] += 1

        await self.db.commit()
        logger.info(f"ETL sync completed: {stats}")
        return stats

    async def _find_candidate_by_hash(self, row_hash: str) -> Optional[CandidateInfo]:
        """Find existing candidate by row hash."""
        result = await self.db.execute(
            select(CandidateInfo).where(CandidateInfo.row_hash == row_hash)
        )
        return result.scalar_one_or_none()

    async def _create_candidate(self, row: Dict[str, Any]) -> CandidateInfo:
        """Create a new candidate from row data."""
        cin = self.generate_cin(
            row.get("offer_release_date"),
            row.get("ref_no")
        )

        candidate = CandidateInfo(
            cin=cin,
            row_hash=row["row_hash"],
            recruiter_name=row.get("recruiter_name"),
            cv_sourced_date=self._parse_date(row.get("cv_sourced_date")),
            jd_published_date=self._parse_date(row.get("jd_published_date")),
            prefix=row.get("prefix"),
            vertical=row.get("vertical"),
            bu=row.get("bu"),
            source_base=row.get("source_base"),
            source=row.get("source"),
            consultant_name=row.get("consultant_name"),
            designation_to_be_printed_on_the_offer_letter=row.get("designation"),
            previous_experience=row.get("previous_experience"),
            grade=row.get("grade"),
            technology=row.get("technology"),
            ref_no=self._parse_date(row.get("ref_no")),
            offer_release_date=self._parse_date(row.get("offer_release_date")),
            expected_doj_wrt_to_np=row.get("expected_doj"),
            month_of_joining=row.get("month_of_joining"),
            current_status=row.get("current_status"),
            personal_email_id=row.get("personal_email_id"),
            contact_number=row.get("contact_number"),
            current_residential_address=row.get("current_address"),
            current_place_of_stay=row.get("place_of_stay"),
            reporting_location=row.get("reporting_location"),
            work_base_location=row.get("work_base_location"),
            np=row.get("np"),
            candidate_name=row.get("candidate_name"),
            candidate_type_id=await self.get_or_create_candidate_type(
                row.get("candidate_type")
            ),
        )

        self.db.add(candidate)
        return candidate

    async def _update_candidate(
        self,
        candidate: CandidateInfo,
        row: Dict[str, Any]
    ) -> None:
        """Update existing candidate with new data."""
        candidate.recruiter_name = row.get("recruiter_name") or candidate.recruiter_name
        candidate.current_status = row.get("current_status") or candidate.current_status
        candidate.updated_on = datetime.utcnow()
        self.db.add(candidate)

    def _parse_date(self, value: Any) -> Optional[datetime]:
        """Parse date value from Excel."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date() if hasattr(value, 'date') else value
        try:
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None